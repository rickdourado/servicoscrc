import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
from datetime import datetime

def format_date(date_str):
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str.split('.')[0], "%Y-%m-%dT%H:%M:%S")
        return dt.strftime("%d/%m/%Y %H:%M")
    except:
        return date_str

def parse_adf(content):
    """Simplifica o Atlassian Document Format para texto plano."""
    if not content:
        return ""
    if isinstance(content, str):
        return content
    
    text = ""
    if isinstance(content, dict):
        # Recursivamente extrair 'text'
        if 'text' in content:
            text += content['text']
        if 'content' in content:
            for item in content['content']:
                text += parse_adf(item)
                if content.get('type') == 'paragraph':
                    text += "\n"
    return text

def generate_excel():
    input_file = '/home/ssdlinux/Documents/dev/servicoscrc/scratch/jira_deep_data.json'
    output_file = '/home/ssdlinux/Documents/dev/servicoscrc/documentacao/relatorio_profundo_jira.xlsx'
    
    if not os.path.exists(input_file):
        print(f"Erro: {input_file} não encontrado.")
        return

    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    issues = data.get('issues', [])
    issue_map = {issue['key']: issue for issue in issues}
    
    # Hierarquia
    parents = []
    for issue in issues:
        parent_info = issue['fields'].get('parent')
        if not parent_info:
            parents.append(issue)

    # Ordenar por chave
    parents.sort(key=lambda x: int(x['key'].split('-')[1]) if '-' in x['key'] else 0)

    # Novo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CRM Deep Report"

    # Definição de Cores
    header_bg = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subtitle_bg = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    parent_bg = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    white_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    std_font = Font(name="Arial", size=10)
    subtitle_font = Font(name="Arial", size=9)
    bold_std_font = Font(name="Arial", size=10, bold=True)

    # 1. Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = ws.cell(row=1, column=1, value="🗺️ MAPA HIERÁRQUICO DO PROJETO PROFUNDO — Entrega por Subtarefas")
    cell.fill = header_bg
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # 2. Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    cell = ws.cell(row=2, column=1, value=f"Exportação Detalhada Jira Profundo | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Inclui Comentários e Changelog")
    cell.fill = subtitle_bg
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # 3. Cabeçalhos
    headers = ['Chave', 'Título / Subtarefa', 'Status', 'Progresso', 'Responsável', 'Prioridade', 'Comentários', 'Histórico de Atividades']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.fill = header_bg
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # 4. Dados
    current_row = 4
    for p in parents:
        # Calcular progresso do bloco
        p_subtasks = p['fields'].get('subtasks', [])
        total_subs = len(p_subtasks)
        done_subs = 0
        for s_ref in p_subtasks:
            s_data = issue_map.get(s_ref['key'])
            if s_data and s_data['fields'].get('status', {}).get('statusCategory', {}).get('key') == 'done':
                done_subs += 1
        
        progress = f"{done_subs}/{total_subs} ({int(done_subs/total_subs*100) if total_subs > 0 else (100 if p['fields'].get('status', {}).get('statusCategory', {}).get('key') == 'done' else 0)}%)"

        # Comentários do Pai
        p_comments = []
        for c in p['fields'].get('comment', {}).get('comments', []):
            author = c.get('author', {}).get('displayName', '?')
            date = format_date(c.get('created'))
            p_comments.append(f"[{date}] {author}: {parse_adf(c.get('body'))}")
        
        # Atividades do Pai
        p_history = []
        for h in p.get('changelog', {}).get('histories', [])[-3:]: # Top 3
            author = h.get('author', {}).get('displayName', '?')
            date = format_date(h.get('created'))
            items = []
            for item in h.get('items', []):
                items.append(f"{item.get('field')}: {item.get('fromString') or 'vazio'} -> {item.get('toString') or 'vazio'}")
            p_history.append(f"({date}) {author}: {'; '.join(items)}")

        row_data = [
            p['key'],
            p['fields'].get('summary'),
            p['fields'].get('status', {}).get('name'),
            progress,
            p['fields'].get('assignee', {}).get('displayName') if p['fields'].get('assignee') else 'Não atribuído',
            p['fields'].get('priority', {}).get('name'),
            "\n".join(p_comments),
            "\n".join(p_history)
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.fill = parent_bg
            cell.font = bold_std_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        current_row += 1
        
        # Subtarefas
        for s_ref in p_subtasks:
            s = issue_map.get(s_ref['key'])
            if not s: continue
            
            s_comments = []
            for c in s['fields'].get('comment', {}).get('comments', []):
                author = c.get('author', {}).get('displayName', '?')
                date = format_date(c.get('created'))
                s_comments.append(f"[{date}] {author}: {parse_adf(c.get('body'))}")
            
            s_history = []
            for h in s.get('changelog', {}).get('histories', [])[-3:]:
                author = h.get('author', {}).get('displayName', '?')
                date = format_date(h.get('created'))
                items = []
                for item in h.get('items', []):
                    items.append(f"{item.get('field')}: {item.get('fromString') or 'vazio'} -> {item.get('toString') or 'vazio'}")
                s_history.append(f"({date}) {author}: {'; '.join(items)}")

            s_row = [
                s['key'],
                "  ↳ " + s['fields'].get('summary'),
                s['fields'].get('status', {}).get('name'),
                "-",
                s['fields'].get('assignee', {}).get('displayName') if s['fields'].get('assignee') else 'Não atribuído',
                s['fields'].get('priority', {}).get('name'),
                "\n".join(s_comments),
                "\n".join(s_history)
            ]
            
            for col, val in enumerate(s_row, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = std_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            current_row += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50

    wb.save(output_file)
    print(f"Excel gerado com sucesso em: {output_file}")

if __name__ == "__main__":
    generate_excel()
