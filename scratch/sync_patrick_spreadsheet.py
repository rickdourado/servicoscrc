import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Mapeamento de Planilha do Patrick -> (Sheet no RevisaoForms.xlsx, Nome do Serviço)
MAPPING = {
    'EstacionamentoIrregular': ('GM-RIO', 'Fiscalização de estacionamento irregular de veículo'),
    'PertubaçãoSossego': ('GM-RIO', 'Fiscalização de perturbação do sossego'),
    'AnimaisSilvestres': ('SMAC', 'Resgate de animais silvestres'),
    'VeiculoAbandonado': ('GFER', 'Remoção de veículo abandonado em via pública'),
    'ObraImovelPrivado': ('SMDU', 'Fiscalização de obras em imóvel privado'),
    'OcupacaoIrregularPub': ('CLF', 'Fiscalização da ocupação irregular de área pública por estabelecimentos comerciais, industriais ou serviços'),
    'AtividadesSemAlvara': ('CLF', 'Fiscalização de atividades econômicas sem alvará'),
    'ImovelRachadura': ('DEFESA CIVIL', 'Vistoria em imóvel com rachadura e infiltração'),
    'AmeacaDesabamento': ('DEFESA CIVIL', 'Vistoria em ameaça de desabamento de estrutura')
}

PATH_PATRICK = 'refs/planilhas/RevisaoForms_PorServico_Patrick.xlsx'
PATH_REVISAO = 'refs/planilhas/RevisaoForms.xlsx'
PATH_OUTPUT = 'refs/planilhas/RevisaoForms_PorServico_Patrick_Sincronizada.xlsx'

def extract_reference_data(path_revisao):
    """
    Extrai um dicionário { (sheet, service): { field: to_be_value } }
    """
    ref_data = {}
    xl = pd.ExcelFile(path_revisao)
    
    for sheet in xl.sheet_names:
        if sheet in ['Campos_Verificação', 'Mails - Pontos Focais']:
            continue
            
        df = pd.read_excel(path_revisao, sheet_name=sheet)
        
        current_service = None
        service_fields = {}
        
        for index, row in df.iterrows():
            first_col = str(row.iloc[0]).strip()
            
            # Detecta se é um cabeçalho de serviço (Geralmente Unnamed: 1 é NaN ou o título ocupa a linha)
            # Na inspeção, vimos que o serviço fica na primeira coluna e as outras são NaN
            if first_col != 'nan' and pd.isna(row.iloc[1]) and pd.isna(row.iloc[3]):
                # Se tínhamos um serviço anterior, salva
                if current_service:
                    ref_data[(sheet, current_service)] = service_fields
                
                current_service = first_col
                service_fields = {}
                continue
            
            if current_service and first_col != 'nan':
                # Pega a coluna D (index 3) que é o "TO BE"
                to_be_val = str(row.iloc[3]).strip()
                # Remove o ':' do nome do campo para bater com a planilha do Patrick
                field_name = first_col.replace(':', '').strip()
                service_fields[field_name] = to_be_val
        
        # Salva o último serviço da aba
        if current_service:
            ref_data[(sheet, current_service)] = service_fields
            
    return ref_data

def sync():
    print("Iniciando extração de dados de referência...")
    ref_data = extract_reference_data(PATH_REVISAO)
    
    print("Abrindo planilha do Patrick...")
    wb = openpyxl.load_workbook(PATH_PATRICK) # use openpyxl directly to preserve formatting
    
    for sheet_name, (ref_sheet, ref_service) in MAPPING.items():
        if sheet_name not in wb.sheetnames:
            print(f"Aviso: Aba {sheet_name} não encontrada no arquivo do Patrick.")
            continue
            
        print(f"Sincronizando {sheet_name} -> {ref_service}...")
        ws = wb[sheet_name]
        
        # Procura os dados de referência
        fields_to_update = ref_data.get((ref_sheet, ref_service))
        if not fields_to_update:
            print(f"ERRO: Não encontrei dados para {ref_service} na aba {ref_sheet}")
            continue
            
        # No Patrick: Coluna A (1) é o Campo, Coluna G (7) é o Cenário Futuro
        # Itera pelas linhas a partir da 4 (onde começam os campos)
        for row in range(4, ws.max_row + 1):
            field_name_cell = ws.cell(row=row, column=1).value
            if not field_name_cell:
                continue
                
            field_name = str(field_name_cell).replace(':', '').strip()
            
            if field_name in fields_to_update:
                new_value = fields_to_update[field_name]
                # Atualiza Column G (7)
                ws.cell(row=row, column=7).value = new_value
                # print(f"  {field_name}: {new_value}")
            else:
                # Tenta match parcial (ex: "Ponto de Referência" vs "Ponto de Referência:")
                match = None
                for k in fields_to_update.keys():
                    if k.lower() in field_name.lower() or field_name.lower() in k.lower():
                        match = k
                        break
                if match:
                    new_value = fields_to_update[match]
                    ws.cell(row=row, column=7).value = new_value
    
    print(f"Salvando resultado em {PATH_OUTPUT}...")
    wb.save(PATH_OUTPUT)
    print("Sincronização concluída com sucesso!")

if __name__ == "__main__":
    sync()
