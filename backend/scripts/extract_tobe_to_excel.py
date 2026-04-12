import os
import glob
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_fields_from_tobe(html_path):
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    fields = []
    
    # Em TO BE forms, os campos estão agrupados no .form-group
    groups = soup.find_all('div', class_='form-group')
    
    for group in groups:
        # Busca o label padrão ou alternativo
        label_tag = group.find('label', class_='form-label')
        if not label_tag:
            label_tag = group.find('label')
            
        if not label_tag:
            continue
            
        # Pega título removendo o asterisco do span.req
        req_span = label_tag.find('span', class_='req')
        is_required = req_span is not None
        required_text = "Sim" if is_required else "Não"
        
        # Limpar titulo (remover o span interno pra não ficar "Título *")
        if req_span:
            req_span.extract()
            
        title = label_tag.get_text(strip=True).strip('*').strip()
        
        status = "Visível" # Default no TO BE
        
        # Input real
        input_tag = group.find(['input', 'select', 'textarea'])
        
        # Grid de checkboxes
        checkbox_grid = group.find('div', class_='checkbox-grid')
        
        description = ""
        field_type = ""
        
        if checkbox_grid:
            field_type = "checkbox (múltiplos)"
            checks = [lbl.get_text(strip=True) for lbl in checkbox_grid.find_all('label')]
            description = f"Valores: {', '.join(checks)}"
        elif input_tag:
            field_type = input_tag.name
            if field_type == 'input':
                field_type = f"input {input_tag.get('type', 'text')}"
            
            if input_tag.has_attr('disabled'):
                status = "Desabilitado"
            elif input_tag.has_attr('readonly'):
                status = "Somente Leitura"
                
            if input_tag.has_attr('placeholder'):
                description = input_tag['placeholder']
                
        # Obter a seção verificando o step-content-section parent
        section = "Geral"
        step_section = group.find_parent('div', class_='step-content-section')
        if step_section:
            title_tag = step_section.find('h2', class_='step-title')
            if title_tag:
                section = title_tag.get_text(strip=True)
                
        fields.append({
            'section': section,
            'title': title,
            'status': status,
            'required': required_text,
            'description': description,
            'type': field_type
        })
        
    # Hack para achar a área de upload de fotos que não é form-group
    photo_area = soup.find(text=lambda x: x and '📸' in x)
    if photo_area:
        step_section = photo_area.find_parent('div', class_='step-content-section')
        section = "Evidências Visuais"
        if step_section:
            title_tag = step_section.find('h2', class_='step-title')
            if title_tag:
                section = title_tag.get_text(strip=True)
                
        fields.append({
            'section': section,
            'title': "Upload de Câmera/Galeria",
            'status': "Visível",
            'required': "Sim (Depende do Fluxo)",
            'description': "Arraste fotos do local para cá (JPG, PNG)",
            'type': "drag-and-drop-area"
        })
        
    return fields

def apply_formatting(ws):
    # Laranja Prefeitura para diferenciar do AS IS (Azul)
    header_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid") 
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    cell_font = Font(name="Segoe UI", size=10, color="333333")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    row_count = ws.max_row
    col_count = ws.max_column
    
    for row in ws.iter_rows(min_row=2, max_row=row_count, max_col=col_count):
        for idx, cell in enumerate(row):
            cell.font = cell_font
            cell.border = thin_border
            
            if idx in [2, 3, 5]:  
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    
    widths = {
        'A': 30, # Seção
        'B': 45, # Título do Campo
        'C': 18, # Status
        'D': 20, # Obrigatório
        'E': 60, # Descrição
        'F': 25  # Tipo HTML
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def process_tobe_wireframes(wireframes_dir, output_excel_path):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Processa somente arquivos com prefixo tobe_
    html_files = glob.glob(os.path.join(wireframes_dir, 'tobe_*.html'))
    
    if not html_files:
        print(f"Nenhum arquivo TO BE HTML encontrado em {wireframes_dir}")
        return

    # Usaremos tobe_wizard_base.html também ou só os especialistas?
    # Filtrar o tobe_wizard_base.html se quisermos apenas os formulários práticos
    
    count = 0
    for html_file in html_files:
        filename = os.path.basename(html_file)
        if filename == 'tobe_wizard_base.html':
            continue
            
        sheet_name = filename.replace('tobe_', '').replace('.html', '')[:31]
        
        fields = extract_fields_from_tobe(html_file)
        
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Seção", "Título do Campo", "Status", "Obrigatório", "Descrição", "Tipo HTML"]
        ws.append(headers)
        
        for f in fields:
            ws.append([f['section'], f['title'], f['status'], f['required'], f['description'], f['type']])
            
        apply_formatting(ws)
        count += 1
            
    if count > 0:
        wb.save(output_excel_path)
        print(f"Extração concluída para {count} wireframes TO BE.")
        print(f"Salvo em: {output_excel_path}")
    else:
        print("Nenhum formulário TO BE real processado.")

if __name__ == "__main__":
    base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    wireframes_path = os.path.join(base_path, 'frontend', 'wireframes')
    output_path = os.path.join(base_path, 'refs', 'planilhas', 'Campos_TO_BE_Mapeados.xlsx')
    
    process_tobe_wireframes(wireframes_path, output_path)
