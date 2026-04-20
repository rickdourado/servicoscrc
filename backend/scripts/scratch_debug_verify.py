import openpyxl
import os

def debug_service(path, service_name_keywords):
    if not os.path.exists(path): return
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- FILE: {os.path.basename(path)} | SHEET: {sheet_name} ---")
        found_service = False
        for row in ws.iter_rows(values_only=True):
            first_cell = str(row[0]) if row[0] else ""
            if any(kw.lower() in first_cell.lower() for kw in service_name_keywords) and not any(row[1:5]):
                print(f"\n[SERVICE FOUND]: {first_cell}")
                found_service = True
                continue
            
            if found_service:
                # If we hit another header or empty row[0], stop or continue
                if not row[0]: continue
                if any(row[1:5]): # Has data
                    print(f"Row: {row}")
                else: # Likely another service header or end of section
                    if found_service and not any(row[1:5]) and row[0]:
                        break

print("=== CLF Verification ===")
debug_service('refs/Formulários Avulsos/Formulários CLF.xlsx', ['ocupação irregular', 'atividades econômicas'])

print("\n=== Defesa Civil Verification ===")
debug_service('refs/Formulários Avulsos/Formulários DEFESA CIVIL.xlsx', ['rachadura', 'desabamento'])

print("\n=== GM Verification ===")
debug_service('refs/Formulários Avulsos/Formulários GM-RIO.xlsx', ['estacionamento', 'perturbação'])
