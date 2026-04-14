import json
from pydantic import BaseModel
from typing import List
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent.parent
COMPARACAO_JSON = BASE_DIR / "backend" / "data" / "comparacao.json"
EXCEL_PATH = str(BASE_DIR / "refs" / "PlanilhaConsolidada.xlsx")
OUTPUT_PATH = str(BASE_DIR / "refs" / "PlanilhaGerada.xlsx")


class Level1Item(BaseModel):
    id: str
    name: str
    level2: List[str]
    source: str


def extract_data() -> List[Level1Item]:
    if not COMPARACAO_JSON.exists():
        print(f"ERRO: Arquivo {COMPARACAO_JSON} não encontrado.")
        return []

    try:
        with open(COMPARACAO_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
            items = data.get("items", [])
            return [Level1Item(**item) for item in items]
    except Exception as e:
        print(f"Erro ao ler JSON de comparacao: {e}")
        return []


def save_new_order(ordered_items: List[Level1Item]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Serviços Consolidados"

    headers = ["Origem", "Nível 1 – Usuário", "Nível 2 – Tipo de Serviço ou Informação"]
    ws.append(headers)

    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="004080", end_color="004080", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_num = 2
    for item in ordered_items:
        records = (
            [(item.source, item.name, n2) for n2 in item.level2]
            if item.level2
            else [(item.source, item.name, "")]
        )
        for record in records:
            ws.append(record)
            ws.cell(row=row_num, column=1).alignment = center_align
            ws.cell(row=row_num, column=2).alignment = left_align
            ws.cell(row=row_num, column=3).alignment = left_align
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

    for col_num in range(1, len(headers) + 1):
        column_target = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=r, column=col_num).value or ""))
            for r in range(1, row_num)
        )
        adjusted_width = min((max_length + 2) * 1.2, 80)
        if col_num == 1 and adjusted_width < 12:
            adjusted_width = 12
        ws.column_dimensions[column_target].width = adjusted_width

    wb.save(OUTPUT_PATH)
    return {"message": "Planilha gerada com sucesso", "path": OUTPUT_PATH}


if __name__ == "__main__":
    d = extract_data()
    print(f"Extraidos {len(d)} elementos globais.")
