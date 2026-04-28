import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Caminhos dos arquivos
file_nivel3 = '/home/ssdlinux/Documents/dev/servicoscrc/refs/planilhas/servicos/nivel3fica.xlsx'
file_acompanhamento = '/home/ssdlinux/Documents/dev/servicoscrc/refs/planilhas/servicos/acompanhamentoservicos.xlsx'
file_output = '/home/ssdlinux/Documents/dev/servicoscrc/refs/planilhas/servicos/servicosconsolidados.xlsx'

# 1. Ler os dados visíveis de nivel3fica.xlsx usando openpyxl (pois pandas ignora o filtro do excel)
wb_nivel3 = load_workbook(file_nivel3)
ws_nivel3 = wb_nivel3.active

visible_titles = []
for idx, row in enumerate(ws_nivel3.iter_rows()):
    row_idx = row[0].row
    # Pula as linhas que estão ocultas pelo filtro
    if ws_nivel3.row_dimensions[row_idx].hidden:
        continue
    # Pula o cabeçalho
    if idx == 0: 
        continue
        
    val = row[0].value
    if val is not None:
        visible_titles.append(val)

# Criar DataFrame apenas com os títulos filtrados (visíveis)
df_nivel3 = pd.DataFrame({'titulo_servico': visible_titles})

# 2. Carregar a planilha de acompanhamento
df_acompanhamento = pd.read_excel(file_acompanhamento, header=1)
df_acompanhamento.columns = df_acompanhamento.columns.str.strip().str.lower()
df_acompanhamento = df_acompanhamento[['titulo_servico', 'descricao_completa']]

# Remover títulos duplicados na planilha de acompanhamento
df_acompanhamento = df_acompanhamento.drop_duplicates(subset=['titulo_servico'], keep='first')

# 3. Fazer o merge (Left Join)
df_consolidado = pd.merge(df_nivel3, df_acompanhamento, on='titulo_servico', how='left')

# 4. Salvar e Formatar
df_consolidado.to_excel(file_output, index=False)

wb_out = load_workbook(file_output)
ws_out = wb_out.active
ws_out.title = "Serviços Consolidados"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
data_alignment = Alignment(vertical="top", wrap_text=True)

for col_num, column_cells in enumerate(ws_out.columns, 1):
    col_letter = get_column_letter(col_num)
    if col_letter == 'A':
        ws_out.column_dimensions[col_letter].width = 50
    elif col_letter == 'B':
        ws_out.column_dimensions[col_letter].width = 100
        
    for idx, cell in enumerate(column_cells):
        if idx == 0:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        else:
            cell.alignment = data_alignment

wb_out.save(file_output)
print(f"Linhas consideradas (visíveis): {len(visible_titles)}")
print(f"Linhas finais no arquivo consolidado: {len(df_consolidado)}")
print(f"Planilha gerada com sucesso em: {file_output}")
