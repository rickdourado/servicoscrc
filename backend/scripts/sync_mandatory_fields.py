import openpyxl
import os
import re

def update_mandatory_fields():
    path_xlsx = r"c:\Users\Patrick Ribeiro\Documents\dev\servicoscrc\refs\planilhas\AS_IS_Atualizada.xlsx"
    dir_wireframes = r"c:\Users\Patrick Ribeiro\Documents\dev\servicoscrc\frontend\wireframes"

    mapping = {
        "irregular_veiculo": "irregular_veiculo.html",
        "alvara_ocupacao": "wireframe_alvara_ocupacao.html",
        "atividades_economicas": "wireframe_atividades_economicas.html",
        "desabamento": "wireframe_desabamento.html",
        "estrutura_imovel": "wireframe_estrutura_imovel.html",
        "fiscalizacao_obras": "wireframe_fiscalizacao_obras.html",
        "meio_ambiente": "wireframe_meio_ambiente.html",
        "sossego": "wireframe_sossego.html",
        "veiculo_abandonado": "wireframe_veiculo_abandonado.html"
    }

    if not os.path.exists(path_xlsx):
        print(f"Erro: Planilha não encontrada em {path_xlsx}")
        return

    wb = openpyxl.load_workbook(path_xlsx)
    
    # Regex para capturar labels com as classes 'req' ou 'required'
    # Captura a classe e o conteúdo do texto do label
    label_regex = re.compile(r'<label[^>]*class=["\'](req|required)["\'][^>]*>(.*?)</label>', re.IGNORECASE | re.DOTALL)

    for sheet_name, html_file in mapping.items():
        if sheet_name not in wb.sheetnames:
            print(f"Aviso: Aba '{sheet_name}' não encontrada no Excel. Pulando...")
            continue
        
        path_html = os.path.join(dir_wireframes, html_file)
        if not os.path.exists(path_html):
            print(f"Aviso: Arquivo HTML '{html_file}' não encontrado. Pulando aba '{sheet_name}'...")
            continue

        print(f"Processando '{sheet_name}' usando '{html_file}'...")
        
        with open(path_html, 'r', encoding='utf-8') as f:
            html_content = f.read()

        # Extrair todos os labels obrigatórios do HTML
        matches = label_regex.findall(html_content)
        mandatory_labels = set()
        for _, text in matches:
            # Limpar o texto: remover tags HTML internas, espaços e ":"
            clean_text = re.sub(r'<[^>]+>', '', text)
            clean_text = clean_text.strip().rstrip(':').lower()
            mandatory_labels.add(clean_text)
        
        print(f" - Encontrados {len(mandatory_labels)} campos obrigatórios no HTML.")

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        try:
            col_titulo = headers.index("Título do Campo") + 1
            col_obrigatorio = headers.index("Obrigatório") + 1
        except ValueError:
            print(f" - [ERRO] Colunas necessárias não encontradas na aba '{sheet_name}'.")
            continue

        updated_count = 0
        for row in range(2, ws.max_row + 1):
            titulo = ws.cell(row=row, column=col_titulo).value
            if titulo:
                clean_titulo = str(titulo).strip().rstrip(':').lower()
                if clean_titulo in mandatory_labels:
                    ws.cell(row=row, column=col_obrigatorio).value = "Sim"
                    updated_count += 1
                else:
                    ws.cell(row=row, column=col_obrigatorio).value = "Não"
        
        print(f" - Aba '{sheet_name}' atualizada com {updated_count} campos 'Sim'.")

    wb.save(path_xlsx)
    print(f"\nPlanilha '{os.path.basename(path_xlsx)}' atualizada com sucesso.")

if __name__ == "__main__":
    update_mandatory_fields()
