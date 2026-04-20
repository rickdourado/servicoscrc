import openpyxl
import os

files = [
    ('CLF', 'refs/Formulários Avulsos/Formulários CLF.xlsx'),
    ('Defesa Civil', 'refs/Formulários Avulsos/Formulários DEFESA CIVIL.xlsx'),
    ('GM', 'refs/Formulários Avulsos/Formulários GM-RIO.xlsx')
]

def analyze_sheet(ws):
    data = {}
    current_service = None
    
    # Based on the structure seen:
    # Col 0: Field Name
    # Col 1: AS IS Visible
    # Col 2: AS IS Mandatory
    # Col 3: TO BE Visible
    # Col 4: TO BE Mandatory
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and not row[1] and not row[2] and not row[3] and not row[4]:
            current_service = row[0].strip()
            data[current_service] = []
        elif current_service and row[0]:
            field_name = row[0].strip()
            to_be_visible = row[3]
            to_be_mandatory = row[4]
            data[current_service].append({
                'field': field_name,
                'visible': to_be_visible,
                'mandatory': to_be_mandatory
            })
    return data

results = {}

for organ, path in files:
    if not os.path.exists(path):
        results[organ] = "File not found"
        continue
    
    wb = openpyxl.load_workbook(path, data_only=True)
    organ_data = {}
    for sheet_name in wb.sheetnames:
        organ_data[sheet_name] = analyze_sheet(wb[sheet_name])
    results[organ] = organ_data

# Print report in a way that I can read it all
import json
print(json.dumps(results, indent=2, ensure_ascii=False))
