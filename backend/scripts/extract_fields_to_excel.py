import os
import glob
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_fields_from_html(html_path):
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    fields = []
    
    rows = soup.find_all(['div'], class_=lambda c: c and ('row' in c or 'two-col' in c))
    
    for row in rows:
        label = row.find('label')
        if not label:
            continue
            
        title = label.get_text(strip=True)
        if not title:
            continue
            
        is_required = 'req' in label.get('class', []) or 'req-blue' in label.get('class', [])
        required_text = "Sim" if is_required else "Não"
        
        desc_div = row.find('div', class_=['hint', 'subinfo'])
        description = desc_div.get_text(strip=True) if desc_div else ""
        
        field_type = ""
        status = "Visível"
        
        input_tag = row.find(['input', 'select', 'textarea'])
        if input_tag:
            field_type = input_tag.name
            if field_type == 'input':
                field_type = f"input {input_tag.get('type', 'text')}"
            
            if input_tag.has_attr('disabled'):
                status = "Desabilitado"
            elif input_tag.has_attr('readonly'):
                status = "Somente Leitura"
                
        static_span = row.find(['span', 'div'], class_=['badge', 'static'])
        if static_span and not input_tag:
            field_type = "badge/static text"
            status = "Informativo"
            if not description:
                description = static_span.get_text(strip=True)
                
        fieldset = row.find_parent('fieldset')
        section = "Geral"
        if fieldset:
            legend = fieldset.find('legend')
            if legend:
                section = legend.get_text(strip=True)
                
        fields.append({
            'section': section,
            'title': title,
            'status': status,
            'required': required_text,
            'description': description,
            'type': field_type
        })
        
    return fields

def apply_formatting(ws):
    header_fill = PatternFill(start_color="2A5788", end_color="2A5788", fill_type="solid") # Azul mais premium
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    cell_font = Font(name="Segoe UI", size=10, color="333333")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Format Headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Format Content Data
    row_count = ws.max_row
    col_count = ws.max_column
    
    for row in ws.iter_rows(min_row=2, max_row=row_count, max_col=col_count):
        for idx, cell in enumerate(row):
            cell.font = cell_font
            cell.border = thin_border
            
            # Centraliza "Status", "Obrigatório", "Tipo HTML"
            if idx in [2, 3, 5]:  
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Auto Filter
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze Panes (Header)
    ws.freeze_panes = "A2"
    
    # Column Widths Fix
    widths = {
        'A': 25, # Seção
        'B': 45, # Título do Campo
        'C': 18, # Status
        'D': 15, # Obrigatório
        'E': 60, # Descrição
        'F': 22  # Tipo HTML
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def process_wireframes(wireframes_dir, output_excel_path):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    html_files = glob.glob(os.path.join(wireframes_dir, '*.html'))
    
    if not html_files:
        print(f"Nenhum arquivo HTML encontrado em {wireframes_dir}")
        return

    for html_file in html_files:
        filename = os.path.basename(html_file)
        sheet_name = filename.replace('wireframe_', '').replace('.html', '')[:31]
        
        fields = extract_fields_from_html(html_file)
        
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Seção", "Título do Campo", "Status", "Obrigatório", "Descrição", "Tipo HTML"]
        ws.append(headers)
        
        for f in fields:
            ws.append([f['section'], f['title'], f['status'], f['required'], f['description'], f['type']])
            
        apply_formatting(ws)
            
    wb.save(output_excel_path)
    print(f"Extração concluída com formatação premium para {len(html_files)} wireframes.")
    print(f"Salvo em: {output_excel_path}")

if __name__ == "__main__":
    base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    wireframes_path = os.path.join(base_path, 'refs', 'wireframes')
    output_path = os.path.join(base_path, 'refs', 'planilhas', 'Campos_AS_IS_Mapeados_Formatado.xlsx')
    
    process_wireframes(wireframes_path, output_path)
