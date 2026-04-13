import openpyxl
import os

def generate_updated_asis():
    path_as_is = r"c:\Users\Patrick Ribeiro\Documents\dev\servicoscrc\refs\planilhas\Campos_AS_IS_Mapeados.xlsx"
    path_revisao = r"c:\Users\Patrick Ribeiro\Documents\dev\servicoscrc\refs\planilhas\Revisão de formulários.xlsx"
    path_output = r"c:\Users\Patrick Ribeiro\Documents\dev\servicoscrc\refs\planilhas\AS_IS_Atualizada.xlsx"

    print(f"Lendo campos irrelevantes de: {os.path.basename(path_revisao)}...")
    wb_revisao = openpyxl.load_workbook(path_revisao, data_only=True)
    # A aba pode estar truncada dependendo do ambiente/leitor, mas usamos o que encontramos na exploração
    aba_relevancia = "Campos sem relevância para o ci" if "Campos sem relevância para o ci" in wb_revisao.sheetnames else "Campos sem relevância para o cidadão"
    
    ws_relevancia = wb_revisao[aba_relevancia]
    irrelevant_fields = set()
    # Dados começam na linha 5 (index 4) baseado na exploração manual anterior
    for row in ws_relevancia.iter_rows(min_row=5, values_only=True):
        if row[0]:
            irrelevant_fields.add(row[0].strip().lower())

    # Adicionando campos solicitados via imagem/texto adicional
    extra_fields = ["categoria", "tipo/subtipo", "prazo"]
    for field in extra_fields:
        irrelevant_fields.add(field)

    print(f"Campos identificados para remoção: {irrelevant_fields}")

    print(f"Processando planilha base: {os.path.basename(path_as_is)}...")
    wb_as_is = openpyxl.load_workbook(path_as_is)
    
    total_removed = 0

    for sheet_name in wb_as_is.sheetnames:
        ws = wb_as_is[sheet_name]
        print(f" - Analisando aba: {sheet_name}")
        
        # Encontrar o índice da coluna "Título do Campo"
        headers = [cell.value for cell in ws[1]]
        try:
            col_index = headers.index("Título do Campo") + 1
        except ValueError:
            print(f"   [AVISO] Coluna 'Título do Campo' não encontrada na aba {sheet_name}. Pulando...")
            continue

        # Iterar de trás para frente para não bagunçar os índices ao deletar
        rows_to_delete = []
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_index).value
            if cell_value:
                # Limpa o valor da célula: remove espaços e ":" no final
                clean_val = str(cell_value).strip().rstrip(":").lower().strip()
                if clean_val in irrelevant_fields:
                    rows_to_delete.append(row_num)
        
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num)
            total_removed += 1
            
        print(f"   Removidas {len(rows_to_delete)} linhas.")

    wb_as_is.save(path_output)
    print(f"\nSucesso! Planilha salva em: {path_output}")
    print(f"Total de campos removidos no consolidado: {total_removed}")

if __name__ == "__main__":
    generate_updated_asis()
