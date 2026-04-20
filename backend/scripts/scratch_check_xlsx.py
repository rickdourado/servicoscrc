import openpyxl
import os

files = [
    'refs/Formulários Avulsos/Formulários CLF.xlsx',
    'refs/Formulários Avulsos/Formulários DEFESA CIVIL.xlsx',
    'refs/Formulários Avulsos/Formulários GM-RIO.xlsx'
]

for file_path in files:
    print(f"\n--- Analyzing: {file_path} ---")
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\nSheet: {sheet_name}")
        ws = wb[sheet_name]
        # Print first 5 rows to see structure
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            print(f"Row {i+1}: {row}")
