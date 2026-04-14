import openpyxl
import json
import os
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent.parent
EXCEL_COMPARACAO_PATH = str(BASE_DIR / 'refs' / 'PlanilhaConsolidada.xlsx')
EXCEL_SERVICOS_PATH = str(BASE_DIR / 'refs' / 'planilhas' / 'ServicosOrganizacao.xlsx')
DATA_DIR = BASE_DIR / 'backend' / 'data'

os.makedirs(DATA_DIR, exist_ok=True)

def convert_comparacao():
    print(f"Lendo {EXCEL_COMPARACAO_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_COMPARACAO_PATH, data_only=True)
    
    sheets_to_extract = [
        {"name_in_excel": "SRGC", "source_name": "SGRC"},
        {"name_in_excel": "Prefrio", "source_name": "Prefrio"}
    ]
    
    result = []
    idx = 0
    
    for sheet_info in sheets_to_extract:
        if sheet_info["name_in_excel"] not in wb.sheetnames:
            print(f"⚠️ Aba {sheet_info['name_in_excel']} não encontrada.")
            continue
            
        ws = wb[sheet_info["name_in_excel"]]
        data_dict = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            nivel1 = row[0]
            nivel2 = row[1]
            
            if not nivel1:
                continue
                
            nivel1 = str(nivel1).strip()
            nivel2 = str(nivel2).strip() if nivel2 else ""
            
            if nivel1 not in data_dict:
                data_dict[nivel1] = []
                
            if nivel2 and nivel2 not in data_dict[nivel1]:
                data_dict[nivel1].append(nivel2)
        
        for n1_name, n2_list in data_dict.items():
            result.append({
                "id": f"n1_{idx}",
                "name": n1_name,
                "level2": n2_list,
                "source": sheet_info["source_name"]
            })
            idx += 1
            
    output_file = DATA_DIR / 'comparacao.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({"items": result}, f, ensure_ascii=False, indent=4)
    print(f"DONE: Comparacao salva em {output_file}")

def convert_servicos():
    print(f"Lendo {EXCEL_SERVICOS_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_SERVICOS_PATH, data_only=True)
    ws = wb.worksheets[1] # "Nova Árvore - Salesforce"

    result = []
    current_n1 = None
    idx = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        n1_val = row[0] if row[0] else None
        n2_val = row[1] if len(row) > 1 and row[1] else None

        if n1_val:
            if current_n1 is not None:
                result.append(current_n1)
            current_n1 = {
                "id": f"n1_{idx}",
                "name": str(n1_val).strip(),
                "children": []
            }
            idx += 1

        if current_n1 and n2_val:
            n2_str = str(n2_val).strip()
            if n2_str.lower() not in ("[subcategorias]", ""):
                current_n1["children"].append({"name": n2_str})

    if current_n1:
        result.append(current_n1)

    output_file = DATA_DIR / 'servicos.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({"items": result}, f, ensure_ascii=False, indent=4)
    print(f"DONE: Hierarquia salva em {output_file}")

if __name__ == "__main__":
    convert_comparacao()
    convert_servicos()
